# -*- coding: utf-8 -*-
"""
Created on Sat Oct 13 19:55:50 2018

@author: Minh Tuấn1
"""

import requests
import lxml.html as lh
import pandas as pd
from bs4 import BeautifulSoup

def crawl(name_company,number):
    ## name = Mã chứng khoán công ty
    ## number = Cần lấy thông tin giao dịch hay cả báo cáo tài chính
    name_company = str(name_company)
    if number == 1:
        page_link = 'http://finance.vietstock.vn/'+ name_company +'/tai-chinh.htm'
        demo = requests.get(page_link)
        demo_data = BeautifulSoup(demo.text, 'html.parser')
        data = demo_data.find_all('table',class_='FFree_Company_Data')

        list_final = []
        for i in range(len(data)):
            value = str(data[i]).split('">')
            list_value = []
            for row in range(0,len(value)):
                first = value[2]
                value_first = value[4]
                second = value[5]
                value_second = value[7]
                third = value[8]
                value_third = value[10]
                fourth = value[11]
                value_fourth = value[13]
                fifth = value[14]
                value_fifth = value[16]
            list_value.append([first, value_first, second, value_second, third, value_third, fourth, value_fourth, fifth, value_fifth])
            value_split = [row.split('\n')[:2] for row in list_value[-1]]
            for i in range(len(value_split)):
                value = value_split[i]
                if i%2 == 0:
                    v = value[1]
                    list_final.append(v)
                elif i%2 != 0:
                    v  = value[0]
                    list_final.append(v)
        name = list_final[::2]
        value = list_final[1::2]

        name = [v.strip(' \n\r') for v in name[:11]]
        value = [v.split('</')[0] for v in value[:11]]
        
        final = [i for i in zip(name,value)]
        date = demo_data.find_all('p',id='date')
        date = str(date).split('">')[1].strip(' \n\r')
        
        file = pd.DataFrame(final)
        file = file.set_index(0).T
        file['Date']= date
        
        value.insert(0,date)
        value.insert(0,'VNM')
        import openpyxl    
        
        filename = 'D:\\Python code\\crawl_function\\crawl.xlsx'
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws = wb[ 'Daily_transfer']
        ws.append(value)
        wb.save(filename)
        
    elif number == 2:
        page_link = 'http://finance.vietstock.vn/'+ name_company +'/tai-chinh.htm'
        demo = requests.get(page_link)
        demo_data = BeautifulSoup(demo.text, 'html.parser')
        data = demo_data.find_all('table',class_='FFree_Company_Data')

        list_final = []
        for i in range(len(data)):
            value = str(data[i]).split('">')
            list_value = []
            for row in range(0,len(value)):
                first = value[2]
                value_first = value[4]
                second = value[5]
                value_second = value[7]
                third = value[8]
                value_third = value[10]
                fourth = value[11]
                value_fourth = value[13]
                fifth = value[14]
                value_fifth = value[16]
            list_value.append([first, value_first, second, value_second, third, value_third, fourth, value_fourth, fifth, value_fifth])
            value_split = [row.split('\n')[:2] for row in list_value[-1]]
            for i in range(len(value_split)):
                value = value_split[i]
                if i%2 == 0:
                    v = value[1]
                    list_final.append(v)
                elif i%2 != 0:
                    v  = value[0]
                    list_final.append(v)
        name = list_final[::2]
        value = list_final[1::2]

        name = [v.strip(' \n\r') for v in name[:11]]
        value = [v.split('</')[0] for v in value[:11]]
        
        final = [i for i in zip(name,value)]
        date = demo_data.find_all('p',id='date')
        date = str(date).split('">')[1].strip(' \n\r')
        
        file = pd.DataFrame(final)
        file = file.set_index(0).T
        file['Date']= date
        
        value.insert(0,date)
        value.insert(0,'VNM')
        import openpyxl    
        
        filename = 'D:\\Python code\\crawl_function\\crawl.xlsx'
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws = wb[ 'Daily_transfer']
        ws.append(value)
        wb.save(filename)
        
        
        
        link = 'http://finance.vietstock.vn/Controls/Report/Data/GetReport.ashx?rptType=BCTT&scode='+ name_company+ '&bizType=1&rptUnit=1000000&rptTermTypeID=2&page=1'
        pg = requests.get(link)
        
        hs = BeautifulSoup(pg.text, 'html.parser')
        
        fl  = hs.find_all('td',class_='BR_colHeader_Time')
        
        time  = str(fl).split('">')
        time_split  = time[-8::2]
        time_final = [row.split('<')[0] for row in time_split] ## Thời gian tính số
        #print(time_final)
        
        title = hs.find_all('td',class_='BR_colHeader_left')
        title = [str(row).split(';">')[1] for row in title]
        title_final = [row.split('</')[0] for row in title]  ## Tiêu đề 3 đoạn lớn 
        #print(title_final)
        
        row_title_first_two = hs.find_all('td', class_='BR_tBody_colName NormalB Padding1')
        row_title_first_two = [str(row).split('2">')[1] for row in row_title_first_two]
        row_title_first_two_final = [row.split('</')[0] for row in row_title_first_two]  ## 2 đoạn đầu nhưng là dữ liệu được bôi đen 
        #print(row_title_first_two_final)
        
        row_title_first_two_small = hs.find_all('td', class_='BR_tBody_colName Normal Padding1')
        row_title_first_two_small = [str(row).split('2">')[1] for row in row_title_first_two_small]
        row_title_first_two_small_final = [row.split('</')[0] for row in row_title_first_two_small] ## Trong 2 đoạn đó nhưng là dữ liệu không bôi đen
        #print(len(row_title_first_two_small_final))
        
        row_title_third = hs.find_all('td',class_='BR_tBody_colName Padding1')
        row_title_third = [str(row).split('1">')[1] for row in row_title_third]
        row_title_third_final = [row.split('</')[0] for row in row_title_third] ## Đoạn cuối cùng 
        #print(row_title_third_final)
        
        number_first = hs.find_all('td',class_='BR_tBody_colValue_mVND NormalB')
        number_first = [str(row).split('B">')[1] for row in number_first]
        number_first_final = [row.split('</')[0] for row in number_first]
        
        
        number_third = hs.find_all('td',class_='BR_tBody_colValue_mVND')
        number_third = [str(row).split('">')[1] for row in number_third[-48:]]
        number_third_final = [row.split('</')[0] for row in number_third] ## Số đoạn 3 Tài chính
        
        
        row_title_first_two_final.extend(row_title_third_final)
        number_first_final.extend(number_third_final)
        
        number_first_final = [number_first_final[i:i+4] for i in range(0,len(number_first_final),4)]
                
        time_final.insert(0,'')
        
        value = [i for i in zip(row_title_first_two_final,number_first_final)]
        
        list_finance = []
        for row in value:
            number = row[1]
            name = row[0]
            number.insert(0,name)
            list_finance.append(number)

        list_finance.insert(0,time_final)
        wb = openpyxl.load_workbook(filename)
        ws1 = wb.active
        ws1 = wb['Finance']
        for row in list_finance:
            row.insert(0,name_company)
            ws1.append(row)
        wb.save(filename)
    return 'Check file'

# =============================================================================
# if __name__ == '__main__':
#     main()       
# =============================================================================
